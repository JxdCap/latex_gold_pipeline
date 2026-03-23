import argparse
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

try:
    from .导出表格 import 双格式写出
except ImportError:
    from 导出表格 import 双格式写出


def 读取输入表格(输入文件: str | Path, 样本条数: Optional[int] = None) -> list[dict]:
    文件路径 = Path(输入文件)
    工作簿 = load_workbook(文件路径, read_only=True, data_only=True)
    工作表 = 工作簿[工作簿.sheetnames[0]]
    结果: list[dict] = []

    for 行号, (单元格值,) in enumerate(
        工作表.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True),
        start=2,
    ):
        结果.append(
            {
                "编号": f"样本{len(结果) + 1:06d}",
                "来源文件": str(文件路径),
                "来源工作表": 工作表.title,
                "来源行号": 行号,
                "原文本": "" if 单元格值 is None else str(单元格值),
            }
        )
        if 样本条数 is not None and len(结果) >= 样本条数:
            break

    工作簿.close()
    return 结果


def 导出导入结果(记录列表: list[dict], 输出目录: str | Path) -> None:
    目录 = Path(输出目录)
    双格式写出(
        记录列表,
        目录 / "原始样本.jsonl",
        目录 / "原始样本.xlsx",
    )


def main() -> None:
    参数解析器 = argparse.ArgumentParser(description="从固定入口 input.xlsx 读取第一个工作表的 A 列。")
    参数解析器.add_argument("--输入文件", default="data/输入表格/input.xlsx")
    参数解析器.add_argument("--样本条数", type=int, default=20)
    参数解析器.add_argument("--输出目录", default="data/原始数据")
    参数 = 参数解析器.parse_args()

    记录列表 = 读取输入表格(参数.输入文件, 参数.样本条数)
    导出导入结果(记录列表, 参数.输出目录)


if __name__ == "__main__":
    main()
